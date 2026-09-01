#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate the Ultimate Master Excel Workbooks with:
1. Every single L, every single D, every single step count listed row-by-row.
2. Training method checkmarks (✓), including INT8 / INT4 Quantization.
3. Steps, Batch Size, Total Batches, Epochs, Total Samples.
4. Datasource Construction details.
5. Optimizer & Hyperparameters.
6. Loss & Reward Function design.
7. Evaluation & Diagnostics protocol.
8. Checkpoint .pt & Params.
9. Mechanistic conclusions.
10. Merged Original README Sheet: 【仓库架构与项目说明_原README】
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_TITLE = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="DDEBF7")
FONT_SECTION = Font(name="Segoe UI", size=11, bold=True, color="1F4E78")
FONT_HEADER = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
FONT_HEADER_CHECK = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
FONT_REGULAR = Font(name="Segoe UI", size=9, color="000000")
FONT_BOLD = Font(name="Segoe UI", size=9, bold=True, color="000000")
FONT_CODE = Font(name="Consolas", size=8.5, color="1F3864")
FONT_CHECK = Font(name="Segoe UI", size=11, bold=True, color="1B5E20")
FONT_EMPTY = Font(name="Segoe UI", size=9, color="D0D0D0")

FILL_NAVY = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER_CFG = PatternFill("solid", fgColor="2F5597")
FILL_HEADER_DATA = PatternFill("solid", fgColor="41719C")
FILL_HEADER_METH = PatternFill("solid", fgColor="1E6B52")
FILL_HEADER_OPT = PatternFill("solid", fgColor="5B4B8A")
FILL_HEADER_RES = PatternFill("solid", fgColor="843C0C")
FILL_HEADER_CONCL = PatternFill("solid", fgColor="4A235A")

FILL_SECTION = PatternFill("solid", fgColor="D9E1F2")
FILL_ZEBRA_LIGHT = PatternFill("solid", fgColor="F9FBFD")
FILL_CHECK_BG = PatternFill("solid", fgColor="E8F5E9")
FILL_SUCCESS = PatternFill("solid", fgColor="E2EFDA")
FILL_ALERT = PatternFill("solid", fgColor="FCE4D6")

THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color='1F3864'),
    bottom=Side(style='medium', color='1F3864')
)

METHOD_COLS = [
    "SFT监督", "CoT竖式", "Plain无CoT", "纯RL_GRPO", "纯RL_REINFORCE", 
    "自问自答", "稀疏采样", "4位加权", "单样本Single", "打包Packed", 
    "MoE专家", "LoRA适配", "Bottleneck低秩", "全局记忆", "LRU遗忘", 
    "ForcedObs真观测", "BFS路径监督", "CrossAttn压缩", "HeapTopM记忆", 
    "DSA注意力", "ALiBi偏置", "RoPE旋转", "INT8动态量化", "INT4低比特"
]

def create_title_block(ws, title_text, subtitle_text, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell1 = ws.cell(1, 1, value=title_text)
    cell1.font = FONT_TITLE
    cell1.fill = FILL_NAVY
    cell1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell2 = ws.cell(2, 1, value=subtitle_text)
    cell2.font = FONT_SUBTITLE
    cell2.fill = FILL_NAVY
    cell2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

def set_master_headers(ws, row_idx=3):
    c1 = ["序号", "实验类别", "实验标识/具体配置", "层数 L", "宽度 d"]
    for idx, h in enumerate(c1, 1):
        cell = ws.cell(row=row_idx, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_CFG
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    c2 = ["训练步数 (Steps)", "批量 (Batch Size)", "总批次数 (Batches)", "等效 Epoch 数", "样本吞吐量 (Samples)"]
    for idx, h in enumerate(c2, 6):
        cell = ws.cell(row=row_idx, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_DATA
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    c3 = ["数据源类型", "操作数位数 (Digits)", "4位偏置比例 (Bias)", "稀疏衰减 (Sparse)", "空格扰动 (Spaces)", "迷宫网格/拓扑"]
    for idx, h in enumerate(c3, 11):
        cell = ws.cell(row=row_idx, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_DATA
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    c4 = ["学习率 LR", "学习率调度 (Schedule)", "预热步数 (Warmup)", "权重衰减 (WD)", "损失/奖励函数设计"]
    for idx, h in enumerate(c4, 17):
        cell = ws.cell(row=row_idx, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_OPT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    m_start = 22
    for idx, m in enumerate(METHOD_COLS, m_start):
        cell = ws.cell(row=row_idx, column=idx, value=m)
        cell.font = FONT_HEADER_CHECK
        cell.fill = FILL_HEADER_METH
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    r_start = m_start + len(METHOD_COLS)
    c6 = [
        "Add1 %", "Add2 %", "Add3 %", "Add4 %", 
        "Sub1 %", "Sub2 %", "Sub3 %", "Sub4 %", 
        "迷宫到达率 %", "撞墙步数", "唯一式 (Unique)", "Loss 损失", "评测协议/诊断设置", "耗时 (s)", "结果与机制归因"
    ]
    for idx, r in enumerate(c6, r_start):
        cell = ws.cell(row=row_idx, column=idx, value=r)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER_CONCL if r == "结果与机制归因" else FILL_HEADER_RES
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = HEADER_BORDER

    ws.row_dimensions[row_idx].height = 32
    ws.freeze_panes = f"D{row_idx+1}"

def write_master_row(ws, row_idx, r_dict, is_zebra=False):
    steps = int(r_dict.get("steps", 0) or 0)
    bs = int(r_dict.get("bs", 0) or 0)
    total_samples = steps * bs if steps and bs else "—"
    epochs = f"{steps * bs / 1000:.1f}" if steps and bs else "—"

    v1 = [
        r_dict.get("id", ""),
        r_dict.get("category", ""),
        r_dict.get("desc", ""),
        r_dict.get("l", ""),
        r_dict.get("d", "")
    ]
    for c, v in enumerate(v1, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = FONT_CODE if c in (1, 4, 5) else FONT_REGULAR
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if c in (1, 4, 5) else "left", vertical="center")
        if is_zebra: cell.fill = FILL_ZEBRA_LIGHT

    v2 = [steps, bs, steps, epochs, total_samples]
    for c, v in enumerate(v2, 6):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = FONT_CODE
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_zebra: cell.fill = FILL_ZEBRA_LIGHT

    v3 = [
        r_dict.get("data_type", "动态生成器 (加减竖式)"),
        r_dict.get("digits", "1-4位"),
        r_dict.get("bias", "0.0"),
        r_dict.get("sparse", "无衰减 (均匀)"),
        r_dict.get("spaces", "0..3 随机"),
        r_dict.get("maze_grid", "—")
    ]
    for c, v in enumerate(v3, 11):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = FONT_REGULAR
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if c in (12, 13, 14, 15) else "left", vertical="center")
        if is_zebra: cell.fill = FILL_ZEBRA_LIGHT

    v4 = [
        r_dict.get("lr", "3e-4"),
        r_dict.get("schedule", "Cosine + Warmup"),
        r_dict.get("warmup", 200),
        r_dict.get("wd", 0.1),
        r_dict.get("loss_fn", "Cross-Entropy (Next Token)")
    ]
    for c, v in enumerate(v4, 17):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.font = FONT_CODE if c in (17, 19, 20) else FONT_REGULAR
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center" if c in (17, 19, 20) else "left", vertical="center")
        if is_zebra: cell.fill = FILL_ZEBRA_LIGHT

    active_methods = set(r_dict.get("methods", []))
    m_start = 22
    for idx, m in enumerate(METHOD_COLS, m_start):
        cell = ws.cell(row=row_idx, column=idx)
        if m in active_methods:
            cell.value = "✓"
            cell.font = FONT_CHECK
            cell.fill = FILL_CHECK_BG
        else:
            cell.value = "—"
            cell.font = FONT_EMPTY
            if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r_start = m_start + len(METHOD_COLS)
    v6 = [
        r_dict.get("add1", "—"),
        r_dict.get("add2", "—"),
        r_dict.get("add3", "—"),
        r_dict.get("add4", "—"),
        r_dict.get("sub1", "—"),
        r_dict.get("sub2", "—"),
        r_dict.get("sub3", "—"),
        r_dict.get("sub4", "—"),
        r_dict.get("reached", "—"),
        r_dict.get("illegal", "—"),
        r_dict.get("unique", "—"),
        r_dict.get("loss", "—"),
        r_dict.get("eval_setup", "固定测试集 n=40"),
        r_dict.get("time_s", "—"),
        r_dict.get("conclusion", "")
    ]
    for idx, v in enumerate(v6, r_start):
        cell = ws.cell(row=row_idx, column=idx, value=v)
        is_concl = (idx == r_start + len(v6) - 1)
        cell.font = FONT_REGULAR if is_concl else FONT_CODE
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="left" if is_concl else "center", vertical="center", wrap_text=is_concl)
        
        if not is_concl:
            try:
                if str(v).endswith("%"):
                    fval = float(str(v).replace("%", ""))
                    if fval >= 90.0: cell.fill = FILL_SUCCESS
                    elif fval == 0.0: cell.fill = FILL_ALERT
                elif is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            except:
                if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
        elif is_zebra:
            cell.fill = FILL_ZEBRA_LIGHT

    ws.row_dimensions[row_idx].height = 24

def adjust_ultimate_widths(ws):
    widths = {
        "A": 10, "B": 15, "C": 26, "D": 8, "E": 8,
        "F": 12, "G": 12, "H": 12, "I": 12, "J": 14,
        "K": 18, "L": 14, "M": 14, "N": 14, "O": 12, "P": 14,
        "Q": 10, "R": 16, "S": 12, "T": 10, "U": 24
    }
    m_start = 22
    for idx in range(m_start, m_start + len(METHOD_COLS)):
        widths[get_column_letter(idx)] = 11
    r_start = m_start + len(METHOD_COLS)
    for idx in range(r_start, r_start + 14):
        widths[get_column_letter(idx)] = 11
    widths[get_column_letter(r_start + 12)] = 18
    widths[get_column_letter(r_start + 14)] = 65

    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def get_full_granular_data():
    from generate_full_granular_excel import build_all_granular_rows
    base_rows = build_all_granular_rows()
    
    for r in base_rows:
        cat = r.get("category", "")
        desc = r.get("desc", "")
        
        if "量化" in cat:
            r["data_type"] = "模型量化 (PTQ 动态/低比特)"
            r["digits"] = "1-4位验证"
            r["bias"] = "0.5"
            r["spaces"] = "0..3 随机"
            r["loss_fn"] = "无重训 (Post-Training Quant)"
            r["eval_setup"] = "量化精度与压缩比基准 (n=40)"
        elif "Plain" in cat or "PLAIN" in r.get("id", ""):
            r["data_type"] = "动态生成器 (无中间草稿/一步输出)"
            r["digits"] = "1-4位 (覆盖单/多位)"
            r["bias"] = "0.0"
            r["spaces"] = "0..3 随机"
            r["loss_fn"] = "Cross-Entropy (Next Token)"
            r["eval_setup"] = "固定测试集 (n=40, 加减1-4位)"
        elif "迷宫" in cat:
            r["data_type"] = "随机完美迷宫 (DFS生成)"
            r["digits"] = "—"
            r["bias"] = "—"
            r["spaces"] = "—"
            r["maze_grid"] = "5x5 ~ 9x9 (4格局部视场)"
            r["loss_fn"] = "GRPO Group Relative Advantage" if "GRPO" in desc else ("BFS Shortest Path CE" if "SFT" in desc else "REINFORCE")
            r["eval_setup"] = "独立评测器 (n=24, 撞墙即停)"
            r["schedule"] = "Cosine + Warmup"
            r["warmup"] = 20
        elif "自问自答" in cat or "RL" in cat:
            r["data_type"] = "自问自答 Selfplay (模型自出题)"
            r["digits"] = "自发探索 (1-4位)"
            r["bias"] = "—"
            r["spaces"] = "1 空格"
            r["loss_fn"] = "Policy Gradient + KL(0.05) + Memory(r/count)" if "记忆" in desc or "MEM" in r["id"] else "Policy Gradient + KL(0.05)"
            r["eval_setup"] = "采样验证 (n=60, 唯一式/正确率)"
            r["schedule"] = "Constant"
            r["warmup"] = 0
            r["lr"] = "1e-5"
        else:
            r["data_type"] = "动态思维链生成器 (竖式逐位进位/借位)"
            r["digits"] = "1-4位 (含最难双4位)"
            r["bias"] = "0.5 (双4位加权)" if "bias" in desc or "0.5" in desc else "0.0"
            r["spaces"] = "0..3 随机"
            r["loss_fn"] = "Cross-Entropy (Next Token Prediction)"
            r["eval_setup"] = "固定测试集 (n=40, 加减1-4位)"
            r["schedule"] = "Cosine + Warmup 200"
            r["warmup"] = 200
            
    return base_rows

def build_readme_sheet(wb):
    ws = wb.create_sheet("仓库架构与项目说明_原README")
    create_title_block(ws, "simpleAI 仓库架构、任务规格与项目背景说明 (原 README 完整归档)",
                       "包含：HF Submodule 拓扑规范、Checkpoints 7大权重元数据、迷宫与加法任务定义、LFS及核心学术结论", 6)
    
    sections = [
        ("一、仓库拓扑结构与 Submodule 规范", [
            ("simpleAI/", "父仓库极简本地工作区，全部模型代码、实验文档与权重挂在 Hugging Face submodule 上。"),
            ("├── additive-rand-transformer/", "HF submodule #1：微型 Transformer 加法与思维链 (CoT) 算术机制探针。"),
            ("├── maze-transformer/", "HF submodule #2：反应式 2D 迷宫导航（纯 RL 零预训练、4格局部视场）。"),
            ("└── .env / .gitmodules", "Submodule 注册与本地密钥配置（.env 永不入库）。")
        ]),
        ("二、核心任务规格定义 (Task Specifications)", [
            ("加法探针 (Additive Probe)", "探究微型 Transformer 在多位加减法中的机制归因：证明 CoT 是草稿纸读取器而非内部推理器。"),
            ("迷宫导航 (Maze Navigation)", "智能体每步仅看四周四格（路 . / 墙 #），输出 U/D/L/R；指向墙/越界直接原地不动。"),
            ("纯 RL 训练机制", "从随机初始化开始，仅依赖稀疏到达终点奖励学习策略，无任何 BFS 最短路径监督/预训练。")
        ]),
        ("三、预训练模型权重元数据 (7 Checkpoints Metadata)", [
            ("l4_d128_cot_bias05_final.pt (10.66 MB)", "4L·128D causal | 机制研究主体（H1–H4），add1-3 100%, add4 35%, sub1-4 97-100%"),
            ("l4_d128_sft_nobias_final.pt (10.66 MB)", "4L·128D causal | RL 基线（add4 起点 27%）"),
            ("l4_d128_grpo_final.pt (3.55 MB)", "4L·128D + GRPO | add4 27%→32%，sub 零回退"),
            ("l4_d128_reinforce_conservative.pt (3.55 MB)", "4L·128D + REINFORCE | 保守策略梯度基线"),
            ("l4_d128_selfplay_anchor.pt (3.56 MB)", "4L·128D 自问自答 | 难度锚定 + 记忆惩罚解决模式坍缩 (唯一式 40/60)"),
            ("l2_d64_attn_dsa.pt (0.65 MB)", "2L·64D DSA 注意力 | 注意力稀疏变体最优 (add1 93.3%)"),
            ("l2_d64_attn_causal.pt (0.65 MB)", "2L·64D causal | 注意力变体基线")
        ]),
        ("四、核心学术结论与机制归因 (Key Scientific Findings)", [
            ("1. CoT 是草稿纸而非推理", "答案生成阶段 88.7% 依赖读取中间和列，对初始操作数敏感度为 0%；草稿被篡改时 100% 顺从错误。"),
            ("2. 数据形态 > 模型容量", "不加 CoT（Plain 模式）哪怕堆叠 10 层多位加法依然恒为 0%。容量无法替代草稿纸。"),
            ("3. 绝不外推 (Zero Extrapolation)", "只训练 1–4 位时，测试 5–7 位准确率全为 0%，模型仅在草稿纸支持的固定槽位内计算。"),
            ("4. 变体与架构表现", "DSA 注意力 > Causal > Linear；GRPO 引导稳定性远超 REINFORCE；LoRA r8 完美阻断灾难性遗忘。")
        ]),
        ("五、权重加载与环境配置代码", [
            ("本地权重加载", "ck = torch.load('checkpoints/xxx.pt'); cfg = TinyGPTConfig(**ck['config']); model = TinyGPT(cfg)"),
            ("HF 镜像端点", "export HF_ENDPOINT=https://hf-mirror.com （解决直连受限网络）"),
            ("Git LFS 运维", "git -C additive-rand-transformer lfs pull （确保落地真文件而非指针）")
        ])
    ]

    r_idx = 4
    for sec_title, items in sections:
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=6)
        c_sec = ws.cell(row=r_idx, column=1, value=sec_title)
        c_sec.font = FONT_SECTION
        c_sec.fill = FILL_SECTION
        c_sec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r_idx].height = 24
        r_idx += 1

        for k, v in items:
            c_k = ws.cell(row=r_idx, column=1, value=k)
            c_k.font = FONT_BOLD
            c_k.border = THIN_BORDER
            c_k.alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=6)
            c_v = ws.cell(row=r_idx, column=2, value=v)
            c_v.font = FONT_REGULAR
            c_v.border = THIN_BORDER
            c_v.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r_idx].height = 22
            r_idx += 1
        r_idx += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 25
    ws.column_dimensions["F"].width = 25

def build_ultimate_workbooks():
    all_rows = get_full_granular_data()
    print(f"Total fully enriched granular rows: {len(all_rows)}")

    out_paths = [
        "/home/luminovoez/agent/ALL_DOCS_EXPERIMENTS_CONFIG_TO_RESULTS.xlsx",
        "/home/luminovoez/agent/additive-rand-transformer/EXPERIMENTS_ALL.xlsx",
        "/home/luminovoez/agent/maze-transformer/EXPERIMENTS_ALL.xlsx"
    ]

    for out_path in out_paths:
        wb = Workbook()

        # Sheet 1: 训练方式_打勾全实验表
        ws1 = wb.active
        ws1.title = "训练方式_打勾全实验表"
        create_title_block(ws1, "simpleAI 实验全景库 — 逐L / 逐d / 逐步骤 / 训练方式打勾全矩阵", 
                           "覆盖 138 项独立实验：包含完整架构、步数、Batch数、Epochs、数据源构建、超参、24项打勾方式(含量化)与细分结果指标", 
                           21 + len(METHOD_COLS) + 15)
        set_master_headers(ws1, 3)

        for idx, row_dict in enumerate(all_rows, 4):
            write_master_row(ws1, idx, row_dict, is_zebra=(idx%2==0))

        adjust_ultimate_widths(ws1)

        # Sheet 2: 层数L_独立扫描
        ws2 = wb.create_sheet("层数L_独立扫描")
        create_title_block(ws2, "模型深度 L 连续扫描全记录 (L=1..10 逐层独立列出)", 
                           "固定 D128, Single CoT 4000步, 观察多位加减法容量相变与溢出瓶颈", 21 + len(METHOD_COLS) + 15)
        set_master_headers(ws2, 3)
        l_rows = [r for r in all_rows if r['category'] == "架构-层数扫描"]
        for idx, r in enumerate(l_rows, 4):
            write_master_row(ws2, idx, r, is_zebra=(idx%2==0))
        adjust_ultimate_widths(ws2)

        # Sheet 3: 宽度d_独立扫描
        ws3 = wb.create_sheet("宽度d_独立扫描")
        create_title_block(ws3, "模型宽度 d 连续扫描全记录 (d=32..512 逐宽独立列出)", 
                           "固定 L2, Single CoT 2000步, 观察通道数对1-4位加减法解锁门槛", 21 + len(METHOD_COLS) + 15)
        set_master_headers(ws3, 3)
        d_rows = [r for r in all_rows if r['category'] == "架构-宽度扫描"]
        for idx, r in enumerate(d_rows, 4):
            write_master_row(ws3, idx, r, is_zebra=(idx%2==0))
        adjust_ultimate_widths(ws3)

        # Sheet 4: 训练步数_增量续训
        ws4 = wb.create_sheet("训练步数_增量续训")
        create_title_block(ws4, "训练步数增量续训记录 (7模型 × +500/+1000/+2000/+4000步 逐项独立列出)", 
                           "验证算术能力与借位借位是否随训练量提升，判定4位加法进位瓶颈本质", 21 + len(METHOD_COLS) + 15)
        set_master_headers(ws4, 3)
        s_rows = [r for r in all_rows if r['category'] == "训练步数-增量续训"]
        for idx, r in enumerate(s_rows, 4):
            write_master_row(ws4, idx, r, is_zebra=(idx%2==0))
        adjust_ultimate_widths(ws4)

        # Sheet 5: 迷宫导航_纯RL实验
        ws5 = wb.create_sheet("迷宫导航_纯RL实验")
        create_title_block(ws5, "反应式 2D 迷宫导航纯 RL 零预训练实验 (每项独立列出)", 
                           "Transformer GRPO vs REINFORCE vs GRU-RNN (60/120/300步) vs 压缩/Heap记忆", 21 + len(METHOD_COLS) + 15)
        set_master_headers(ws5, 3)
        m_rows = [r for r in all_rows if r['category'] == "迷宫导航-纯RL"]
        for idx, r in enumerate(m_rows, 4):
            write_master_row(ws5, idx, r, is_zebra=(idx%2==0))
        adjust_ultimate_widths(ws5)

        # Sheet 6: 仓库架构与项目说明_原README (Merged original README)
        build_readme_sheet(wb)

        wb.save(out_path)
        print(f"Ultimate Workbook with Merged README saved to: {out_path}")

if __name__ == "__main__":
    build_ultimate_workbooks()
