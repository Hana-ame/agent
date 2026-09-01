#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate all JSON configs strictly named '{序号}_{描述}.json'
where {序号} exactly matches the '序号' column in the Excel workbooks.
"""

import os
import re
import json
import shutil
import openpyxl

def sanitize(name):
    s = re.sub(r'[^a-zA-Z0-9_\-]+', '_', str(name)).strip('_')
    return s.lower()

# ------------------------------------------------------------------------------
# 1. ADDITIVE TRANSFORMER CONFIGS
# ------------------------------------------------------------------------------
def generate_additive_configs():
    wb_p = "/home/luminovoez/agent/additive-rand-transformer/EXPERIMENTS_ALL.xlsx"
    wb = openpyxl.load_workbook(wb_p)
    ws = wb["加法算术探针全实验表"]
    
    cfg_dir = "/home/luminovoez/agent/additive-rand-transformer/configs"
    if os.path.exists(cfg_dir):
        shutil.rmtree(cfg_dir)
    os.makedirs(cfg_dir, exist_ok=True)
    
    count = 0
    for r in range(4, ws.max_row + 1):
        exp_id = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        l = ws.cell(r, 4).value
        d = ws.cell(r, 5).value
        steps = ws.cell(r, 6).value
        bs = ws.cell(r, 7).value
        
        if not exp_id or not desc:
            continue
            
        # Build JSON dictionary
        clean_desc = sanitize(desc)
        filename = f"{exp_id}_{clean_desc}.json"
        
        is_cot = "CoT" in str(desc) or "cot" in str(exp_id).lower() or "L-" in str(exp_id) or "D-" in str(exp_id)
        if "Plain" in str(cat) or "PLAIN" in str(exp_id):
            is_cot = False
            
        cfg = {
            "id": exp_id,
            "category": cat,
            "description": desc,
            "layers": int(l) if str(l).isdigit() else 2,
            "d": int(d) if str(d).isdigit() else 64,
            "heads": 4,
            "steps": int(steps) if str(steps).isdigit() else 4000,
            "batch_size": int(bs) if str(bs).isdigit() else 32,
            "lr": 3e-4,
            "wd": 0.1,
            "warmup": 200,
            "datasource": {
                "type": "cot" if is_cot else "plain",
                "max_digits": 4,
                "bias": 0.5 if ("bias" in str(desc).lower() or "0.5" in str(desc)) else 0.0,
                "max_spaces": 3,
                "single": True
            }
        }
        
        # Specific overrides
        if "NHEAD" in exp_id:
            h_match = re.search(r'n_head\s*=\s*(\d+)', desc)
            if h_match:
                cfg["heads"] = int(h_match.group(1))
        elif "DP-" in exp_id:
            dp_match = re.search(r'dropout\s*=\s*([\d\.]+)', desc)
            if dp_match:
                cfg["dropout"] = float(dp_match.group(1))
        elif "ATTN-" in exp_id:
            attn_match = re.search(r'attn_type\s*=\s*([a-zA-Z0-9]+)', desc)
            if attn_match:
                cfg["attn_type"] = attn_match.group(1).lower()
        elif "MOE-" in exp_id:
            cfg["n_experts"] = 4
            cfg["moe_topk"] = 2
            if "n_experts" in desc:
                m_exp = re.search(r'n_experts\s*=\s*(\d+)', desc)
                if m_exp: cfg["n_experts"] = int(m_exp.group(1))
            if "moe_topk" in desc:
                m_topk = re.search(r'moe_topk\s*=\s*(\d+)', desc)
                if m_topk: cfg["moe_topk"] = int(m_topk.group(1))
            if "moe_aux" in desc:
                m_aux = re.search(r'moe_aux\s*=\s*([\d\.]+)', desc)
                if m_aux: cfg["moe_aux"] = float(m_aux.group(1))
        elif "BN-" in exp_id:
            if "bottleneck" in desc:
                m_bn = re.search(r'bottleneck\s*=\s*(\d+)', desc)
                if m_bn: cfg["bottleneck"] = int(m_bn.group(1))
            if "pos" in desc:
                m_pos = re.search(r'pos\s*=\s*([a-zA-Z0-9_]+)', desc)
                if m_pos: cfg["bottleneck_pos"] = m_pos.group(1)
        elif "RL-" in exp_id:
            cfg["task"] = "selfplay"
            cfg["lr"] = 1e-5
            cfg["steps"] = 150
            if "MEM" in exp_id:
                m_mem = re.search(r'memory_bonus\s*=\s*([\d\.]+)', desc)
                if m_mem: cfg["memory_bonus"] = float(m_mem.group(1))
                
        out_file = os.path.join(cfg_dir, filename)
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        count += 1
        
    print(f"✓ Generated {count} additive configs in: {cfg_dir}")

# ------------------------------------------------------------------------------
# 2. MAZE TRANSFORMER CONFIGS
# ------------------------------------------------------------------------------
def generate_maze_configs():
    wb_p = "/home/luminovoez/agent/maze-transformer/EXPERIMENTS_ALL.xlsx"
    wb = openpyxl.load_workbook(wb_p)
    ws = wb["迷宫纯RL与导航实验全景表"]
    
    cfg_dir = "/home/luminovoez/agent/maze-transformer/configs"
    if os.path.exists(cfg_dir):
        if os.path.islink(cfg_dir): os.unlink(cfg_dir)
        else: shutil.rmtree(cfg_dir)
    os.makedirs(cfg_dir, exist_ok=True)
    
    count = 0
    for r in range(4, ws.max_row + 1):
        exp_id = ws.cell(r, 1).value
        cat = ws.cell(r, 2).value
        desc = ws.cell(r, 3).value
        l = ws.cell(r, 4).value
        d = ws.cell(r, 5).value
        h = ws.cell(r, 6).value
        steps = ws.cell(r, 7).value
        bs = ws.cell(r, 8).value
        
        if not exp_id or not desc:
            continue
            
        clean_desc = sanitize(desc)
        filename = f"{exp_id}_{clean_desc}.json"
        
        cfg = {
            "id": exp_id,
            "category": cat,
            "description": desc,
            "layers": int(l) if str(l).isdigit() else 2,
            "d": int(d) if str(d).isdigit() else 64,
            "heads": int(h) if str(h).isdigit() else 4,
            "steps": int(steps) if str(steps).isdigit() else 120,
            "batch_size": int(bs) if str(bs).isdigit() else 6,
            "lr": 3e-4,
            "min_size": 5,
            "max_size": 9,
            "single": True,
            "datasource": {
                "type": "random_perfect_maze",
                "observation": "forced_obs_4cell",
                "reward": "sparse_goal_reach"
            }
        }
        
        if "RNN" in exp_id:
            cfg["model_type"] = "rnn"
            cfg["layers"] = 1
            cfg["d"] = 128
            cfg["heads"] = 1
        elif "SFT" in exp_id:
            cfg["task"] = "sft_bfs"
            cfg["steps"] = 2000
        elif "CTX" in exp_id:
            cfg["use_context_compression"] = True
        elif "HEAP" in exp_id:
            cfg["use_topm_heap"] = True
            
        out_file = os.path.join(cfg_dir, filename)
        with open(out_file, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2, ensure_ascii=False)
        count += 1
        
    print(f"✓ Generated {count} maze configs in: {cfg_dir}")

if __name__ == "__main__":
    generate_additive_configs()
    generate_maze_configs()
