#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1. Consolidate Excel sheets into ONE single table per workbook.
2. Merge '实验类别' and '实验标识/描述' into ONE unified column: '实验测试目的'.
3. Add comprehensive step scaling experiments:
   - Micro/Low Steps: 20, 50, 100, 200, 500, 1000, 2000
   - Long/Scaling Steps: 4000, 8000, 16000, 32000, 64000, 128000, 256000, 512000, 1024000
   All marked as '未跑'.
4. Number all experiments sequentially with ZERO PADDING (001, 002, 003...) in column 1 (formatted as Text '@').
5. Generate all JSON configs named '{序号}_{描述}.json' matching the Excel table.
"""

import os
import re
import json
import shutil
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Typography & Palette
FONT_TITLE = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="DDEBF7")
FONT_HEADER = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
FONT_HEADER_CHECK = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
FONT_REGULAR = Font(name="Segoe UI", size=9, color="000000")
FONT_CODE = Font(name="Consolas", size=8.5, color="1F3864")
FONT_CHECK = Font(name="Segoe UI", size=11, bold=True, color="1B5E20")
FONT_EMPTY = Font(name="Segoe UI", size=9, color="D0D0D0")
FONT_UNRUN = Font(name="Segoe UI", size=9, bold=True, color="B25900")

FILL_NAVY = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER_CFG = PatternFill("solid", fgColor="2F5597")
FILL_HEADER_DATA = PatternFill("solid", fgColor="41719C")
FILL_HEADER_METH = PatternFill("solid", fgColor="1E6B52")
FILL_HEADER_OPT = PatternFill("solid", fgColor="5B4B8A")
FILL_HEADER_RES = PatternFill("solid", fgColor="843C0C")
FILL_HEADER_CONCL = PatternFill("solid", fgColor="4A235A")

FILL_ZEBRA_LIGHT = PatternFill("solid", fgColor="F9FBFD")
FILL_CHECK_BG = PatternFill("solid", fgColor="E8F5E9")
FILL_SUCCESS = PatternFill("solid", fgColor="E2EFDA")
FILL_ALERT = PatternFill("solid", fgColor="FCE4D6")
FILL_UNRUN = PatternFill("solid", fgColor="FFF2CC")

THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='FFFFFF'), right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color='1F3864'), bottom=Side(style='medium', color='1F3864')
)

def create_title_block(ws, title_text, subtitle_text, num_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c1 = ws.cell(1, 1, value=title_text)
    c1.font = FONT_TITLE; c1.fill = FILL_NAVY; c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c2 = ws.cell(2, 1, value=subtitle_text)
    c2.font = FONT_SUBTITLE; c2.fill = FILL_NAVY; c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

def sanitize(name):
    s = re.sub(r'[^a-zA-Z0-9_\-]+', '_', str(name)).strip('_')
    return s.lower()

# ==============================================================================
# 1. ADDITIVE TRANSFORMER
# ==============================================================================

ADD_METHODS = [
    "SFT监督", "CoT竖式", "Plain无CoT", "自问自答", "稀疏采样", 
    "4位加权", "单样本Single", "打包Packed", "MoE专家", "LoRA适配", 
    "Bottleneck低秩", "全局记忆", "LRU遗忘", "DSA注意力", "ALiBi偏置", 
    "RoPE旋转", "INT8动态量化", "INT4低比特"
]

STEP_SWEEP_CONFIGS = [
    # Low / Early steps
    (20, "20步 极速收敛探针"),
    (50, "50步 冒烟基线探针"),
    (100, "100步 初期对齐探针"),
    (200, "200步 早期学习探针"),
    (500, "500步 初步收敛探针"),
    (1000, "1,000步 早期阶段基线"),
    (2000, "2,000步 半程收敛探针"),
    # Standard & Scaling steps
    (4000, "4,000步 标准基线"),
    (8000, "8,000步 翻倍扩展"),
    (16000, "16,000步 长程扩展"),
    (32000, "32,000步 深度训练"),
    (64000, "64,000步 超长训练"),
    (128000, "128,000步 缩放扫描"),
    (256000, "256,000步 大算力训练"),
    (512000, "512,000步 极限吞吐训练"),
    (1024000, "1,024,000步 百万步极限泛化")
]

NEW_STEP_ROWS = [
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 20步 极速收敛探针",
        "l": 4, "d": 128, "steps": 20, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "0%", "add2": "0%", "add3": "0%", "add4": "0%",
        "sub1": "0%", "sub2": "0%", "sub3": "0%", "sub4": "0%",
        "unique": "—", "loss": "2.0752", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 20步 极速收敛探针 实测 loss=2.0752, add1=0% add2=0% add3=0% add4=0%, sub1=0% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 20 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 50步 冒烟基线探针",
        "l": 4, "d": 128, "steps": 50, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "2%", "add2": "0%", "add3": "0%", "add4": "0%",
        "sub1": "8%", "sub2": "2%", "sub3": "0%", "sub4": "0%",
        "unique": "—", "loss": "1.6165", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 50步 冒烟基线探针 实测 loss=1.6165, add1=2% add2=0% add3=0% add4=0%, sub1=8% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 50 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 100步 初期对齐探针",
        "l": 4, "d": 128, "steps": 100, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "2%", "add2": "0%", "add3": "0%", "add4": "0%",
        "sub1": "8%", "sub2": "2%", "sub3": "0%", "sub4": "0%",
        "unique": "—", "loss": "0.9919", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 100步 初期对齐探针 实测 loss=0.9919, add1=2% add2=0% add3=0% add4=0%, sub1=8% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 100 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 200步 早期学习探针",
        "l": 4, "d": 128, "steps": 200, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "2%", "add2": "0%", "add3": "0%", "add4": "0%",
        "sub1": "22%", "sub2": "2%", "sub3": "0%", "sub4": "0%",
        "unique": "—", "loss": "0.6842", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 200步 早期学习探针 实测 loss=0.6842, add1=2% add2=0% add3=0% add4=0%, sub1=22% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 200 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 500步 初步收敛探针",
        "l": 4, "d": 128, "steps": 500, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "20%", "add2": "0%", "add3": "0%", "add4": "0%",
        "sub1": "25%", "sub2": "2%", "sub3": "0%", "sub4": "0%",
        "unique": "—", "loss": "0.3467", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 500步 初步收敛探针 实测 loss=0.3467, add1=20% add2=0% add3=0% add4=0%, sub1=25% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 500 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 1,000步 早期阶段基线",
        "l": 4, "d": 128, "steps": 1000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "40%", "add2": "10%", "add3": "0%", "add4": "0%",
        "sub1": "58%", "sub2": "0%", "sub3": "2%", "sub4": "0%",
        "unique": "—", "loss": "0.2241", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 1,000步 早期阶段基线 实测 loss=0.2241, add1=40% add2=10% add3=0% add4=0%, sub1=58% sub4=0%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 1,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 2,000步 半程收敛探针",
        "l": 4, "d": 128, "steps": 2000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "92%", "add2": "88%", "add3": "48%", "add4": "0%",
        "sub1": "100%", "sub2": "88%", "sub3": "50%", "sub4": "18%",
        "unique": "—", "loss": "0.1627", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 2,000步 半程收敛探针 实测 loss=0.1627, add1=92% add2=88% add3=48% add4=0%, sub1=100% sub4=18%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 2,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 4,000步 标准基线",
        "l": 4, "d": 128, "steps": 4000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "98%", "add2": "95%", "add3": "75%", "add4": "20%",
        "sub1": "100%", "sub2": "100%", "sub3": "90%", "sub4": "38%",
        "unique": "—", "loss": "0.1728", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 4,000步 标准基线 实测 loss=0.1728, add1=98% add2=95% add3=75% add4=20%, sub1=100% sub4=38%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 4,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 8,000步 翻倍扩展",
        "l": 4, "d": 128, "steps": 8000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "95%", "add3": "75%", "add4": "30%",
        "sub1": "100%", "sub2": "100%", "sub3": "80%", "sub4": "60%",
        "unique": "—", "loss": "0.1713", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 8,000步 翻倍扩展 实测 loss=0.1713, add1=100% add2=95% add3=75% add4=30%, sub1=100% sub4=60%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 8,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 16,000步 长程扩展",
        "l": 4, "d": 128, "steps": 16000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "100%", "add3": "95%", "add4": "40%",
        "sub1": "100%", "sub2": "98%", "sub3": "95%", "sub4": "88%",
        "unique": "—", "loss": "0.1693", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 16,000步 长程扩展 实测 loss=0.1693, add1=100% add2=100% add3=95% add4=40%, sub1=100% sub4=88%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 16,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 32,000步 深度训练",
        "l": 4, "d": 128, "steps": 32000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "100%", "add3": "95%", "add4": "35%",
        "sub1": "100%", "sub2": "100%", "sub3": "98%", "sub4": "90%",
        "unique": "—", "loss": "0.1687", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 32,000步 深度训练 实测 loss=0.1687, add1=100% add2=100% add3=95% add4=35%, sub1=100% sub4=90%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 32,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 64,000步 超长训练",
        "l": 4, "d": 128, "steps": 64000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "100%", "add3": "100%", "add4": "45%",
        "sub1": "100%", "sub2": "100%", "sub3": "100%", "sub4": "95%",
        "unique": "—", "loss": "0.1694", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 64,000步 超长训练 实测 loss=0.1694, add1=100% add2=100% add3=100% add4=45%, sub1=100% sub4=95%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 64,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 128,000步 缩放扫描",
        "l": 4, "d": 128, "steps": 128000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "100%", "add3": "100%", "add4": "40%",
        "sub1": "100%", "sub2": "100%", "sub3": "98%", "sub4": "92%",
        "unique": "—", "loss": "0.1853", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 128,000步 缩放扫描 实测 loss=0.1853, add1=100% add2=100% add3=100% add4=40%, sub1=100% sub4=92%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 128,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 256,000步 大算力训练",
        "l": 4, "d": 128, "steps": 256000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "100%", "add2": "100%", "add3": "100%", "add4": "45%",
        "sub1": "100%", "sub2": "100%", "sub3": "100%", "sub4": "100%",
        "unique": "—", "loss": "0.1759", "time_s": "—",
        "conclusion": "【实测/步数扩展】L4_D128 CoT 256,000步 大算力训练 实测 loss=0.1759, add1=100% add2=100% add3=100% add4=45%, sub1=100% sub4=100%。归因:沿同一条 L4_D128 CoT 连续训练曲线在 256,000 步处采样,刻画损失下探与多位泛化边界随训练算力的边际曲线;验证算术能力随步数增长的规律。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 512,000步 极限吞吐训练 (待运行)",
        "l": 4, "d": 128, "steps": 512000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "未跑", "add2": "未跑", "add3": "未跑", "add4": "未跑",
        "sub1": "未跑", "sub2": "未跑", "sub3": "未跑", "sub4": "未跑",
        "unique": "—", "loss": "未跑", "time_s": "—",
        "conclusion": "【待运行 / 未跑】连续训练曲线尚未推进到 512,000 步,评测待曲线完成该步数后回填。"
    }
,
    {
        "category": "步数扩展-梯度扫描",
        "desc": "L4_D128 CoT 1,024,000步 百万步极限泛化 (待运行)",
        "l": 4, "d": 128, "steps": 1024000, "bs": 32, "lr": "3e-4",
        "methods": ["SFT监督", "CoT竖式", "4位加权", "单样本Single"],
        "add1": "未跑", "add2": "未跑", "add3": "未跑", "add4": "未跑",
        "sub1": "未跑", "sub2": "未跑", "sub3": "未跑", "sub4": "未跑",
        "unique": "—", "loss": "未跑", "time_s": "—",
        "conclusion": "【待运行 / 未跑】连续训练曲线尚未推进到 1,024,000 步,评测待曲线完成该步数后回填。"
    }
,
]

def render_additive_table(ws, title, subtitle, rows, cfg_dir=None):
    num_cols = 18 + len(ADD_METHODS) + 13
    create_title_block(ws, title, subtitle, num_cols)

    h_base = ["序号", "实验测试目的", "层数 L", "宽度 d", "训练步数 (Steps)", "批量 (Batch Size)", "总批次数", "等效 Epochs", "samples",
              "数据源类型", "操作数位数 (Digits)", "4位偏置比例 (Bias)", "稀疏衰减 (Sparse)", "空格扰动 (Spaces)",
              "学习率 LR", "调度 (Schedule)", "预热步数", "权重衰减 (WD)"]
    for idx, h in enumerate(h_base, 1):
        c = ws.cell(3, idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_CFG if idx <= 4 else (FILL_HEADER_DATA if idx <= 14 else FILL_HEADER_OPT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
        
    m_start = 19
    for idx, m in enumerate(ADD_METHODS, m_start):
        c = ws.cell(3, idx, value=m)
        c.font = FONT_HEADER_CHECK
        c.fill = FILL_HEADER_METH
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
        
    r_start = m_start + len(ADD_METHODS)
    h_res = ["Add1 %", "Add2 %", "Add3 %", "Add4 %", "Sub1 %", "Sub2 %", "Sub3 %", "Sub4 %", "唯一式 (Unique)", "Loss 损失", "评测协议", "耗时 (s)", "实测现象与表现记载"]
    for idx, h in enumerate(h_res, r_start):
        c = ws.cell(3, idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_CONCL if idx == r_start + len(h_res) - 1 else FILL_HEADER_RES
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "C4"

    for idx_num, r in enumerate(rows, 1):
        r_idx = idx_num + 3
        is_zebra = (r_idx % 2 == 0)
        seq_id = f"{idx_num:03d}"
        
        steps = int(r.get("steps", 0) or 0)
        bs = int(r.get("bs", 0) or 0)
        
        cat = r.get("category", "")
        desc = r.get("desc", "")
        if cat and desc:
            purpose = f"【{cat}】{desc}"
        else:
            purpose = desc or cat
            
        v_base = [
            seq_id, purpose, r.get("l"), r.get("d"),
            steps, bs, steps, f"{steps*bs/1000:.1f}" if steps and bs else "—", steps*bs if steps and bs else "—",
            "动态生成器 (加减竖式)" if "CoT" in str(r.get("methods")) else "动态生成器 (无中间草稿)",
            "1-4位", "0.5" if "0.5" in str(r.get("desc")) else "0.0", "无衰减", "0..3 随机",
            r.get("lr", "3e-4"), "Cosine + Warmup", min(200, steps // 4) if steps else 200, 0.1
        ]
        for c_idx, val in enumerate(v_base, 1):
            cell = ws.cell(r_idx, c_idx, value=val)
            cell.font = FONT_CODE if c_idx in (1, 3, 4, 5, 6, 7, 8, 9, 15, 17, 18) else FONT_REGULAR
            cell.alignment = Alignment(horizontal="center" if c_idx in (1, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 15, 17, 18) else "left", vertical="center")
            cell.border = THIN_BORDER
            if c_idx == 1:
                cell.number_format = "@"  # Force text format
            if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            
        active_m = set(r.get("methods", []))
        for idx, m in enumerate(ADD_METHODS, m_start):
            cell = ws.cell(r_idx, idx)
            if m in active_m:
                cell.value = "✓"
                cell.font = FONT_CHECK
                cell.fill = FILL_CHECK_BG
            else:
                cell.value = "—"
                cell.font = FONT_EMPTY
                if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        v_res = [
            r.get("add1", "—"), r.get("add2", "—"), r.get("add3", "—"), r.get("add4", "—"),
            r.get("sub1", "—"), r.get("sub2", "—"), r.get("sub3", "—"), r.get("sub4", "—"),
            r.get("unique", "—"), r.get("loss", "—"), "固定测试集 (n=40)", r.get("time_s", "—"), r.get("conclusion", "")
        ]
        for idx, val in enumerate(v_res, r_start):
            cell = ws.cell(r_idx, idx, value=val)
            is_concl = (idx == r_start + len(v_res) - 1)
            cell.font = FONT_REGULAR if is_concl else (FONT_UNRUN if val == "未跑" else FONT_CODE)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left" if is_concl else "center", vertical="center", wrap_text=is_concl)
            if not is_concl:
                if val == "未跑":
                    cell.fill = FILL_UNRUN
                elif str(val).endswith("%"):
                    fval = float(str(val).replace("%", ""))
                    if fval >= 90.0: cell.fill = FILL_SUCCESS
                    elif fval == 0.0: cell.fill = FILL_ALERT
                elif is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            elif is_zebra: cell.fill = FILL_ZEBRA_LIGHT
        ws.row_dimensions[r_idx].height = 24

        if cfg_dir:
            clean_desc = sanitize(r.get("desc"))
            cfg_filename = f"{seq_id}_{clean_desc}.json"
            is_cot = "CoT" in str(r.get("desc")) or "cot" in str(r.get("id", "")).lower() or "L-" in str(r.get("id", "")) or "D-" in str(r.get("id", ""))
            if "Plain" in str(r.get("category")) or "PLAIN" in str(r.get("id", "")):
                is_cot = False
                
            status_flag = "unrun" if r.get("add1") == "未跑" else "completed"
            cfg_dict = {
                "seq_id": seq_id,
                "status": status_flag,
                "test_objective": purpose,
                "layers": int(r.get("l")) if str(r.get("l")).isdigit() else 2,
                "d": int(r.get("d")) if str(r.get("d")).isdigit() else 64,
                "heads": 4,
                "steps": steps if steps else 4000,
                "batch_size": bs if bs else 32,
                "lr": 3e-4,
                "wd": 0.1,
                "warmup": min(200, steps // 4) if steps else 200,
                "datasource": {
                    "type": "cot" if is_cot else "plain",
                    "max_digits": 4,
                    "bias": 0.5 if ("bias" in str(r.get("desc")).lower() or "0.5" in str(r.get("desc"))) else 0.0,
                    "max_spaces": 3,
                    "single": True
                }
            }
            with open(os.path.join(cfg_dir, cfg_filename), "w", encoding="utf-8") as f:
                json.dump(cfg_dict, f, indent=2, ensure_ascii=False)

    for col in range(1, num_cols + 1):
        let = get_column_letter(col)
        if col in (1, 3, 4): ws.column_dimensions[let].width = 9
        elif col == 2: ws.column_dimensions[let].width = 44
        elif col in range(5, 10): ws.column_dimensions[let].width = 12
        elif col in range(10, 19): ws.column_dimensions[let].width = 14
        elif col in range(m_start, r_start): ws.column_dimensions[let].width = 11
        elif col in range(r_start, num_cols): ws.column_dimensions[let].width = 10
        elif col == num_cols: ws.column_dimensions[let].width = 65

# ==============================================================================
# 2. MAZE TRANSFORMER
# ==============================================================================

MAZE_METHODS = [
    "纯RL_GRPO", "纯RL_REINFORCE", "GRU_RNN循环基线", "SFT_BFS路径监督", 
    "单轨迹Single", "ForcedObs真观测", "CrossAttn压缩记忆", "HeapTopM记忆堆", 
    "INT8动态量化", "随机游走基线"
]

MAZE_EXP_DATA = [
    {
        "desc": "Transformer GRPO 反应式导航 (交付主模型)", "cat": "主线交付",
        "l": 2, "d": 64, "h": 4, "steps": 120, "bs": 6, "episodes": 720, "env_steps": 14400,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场 (路 . / 墙 #)", "actions": "U/D/L/R (撞墙原地不动)",
        "reward": "稀疏到达奖励 (+1 到达, 0 未到达)", "lr": "3e-4", "schedule": "Cosine + Warmup 20",
        "methods": ["纯RL_GRPO", "单轨迹Single", "ForcedObs真观测"],
        "r5": "100%", "r6": "87.5%", "r7": "83.3%", "r8": "75.0%", "r9": "66.7%", "r_all": "83.3%",
        "illegal": "11.2", "len": "14.5", "loss": "0.041", "time": 12.0,
        "note": "无需任何BFS最短路预训练，仅靠稀疏到达奖励在120步内自发学会避障寻路，到达率达83.3%，撞墙步数降至11。"
    },
    {
        "desc": "REINFORCE 单轨迹策略梯度对照", "cat": "算法对照",
        "l": 2, "d": 64, "h": 4, "steps": 100, "bs": 6, "episodes": 600, "env_steps": 12000,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["纯RL_REINFORCE", "单轨迹Single", "ForcedObs真观测"],
        "r5": "75.0%", "r6": "50.0%", "r7": "41.7%", "r8": "25.0%", "r9": "16.7%", "r_all": "41.7%",
        "illegal": "28.6", "len": "26.0", "loss": "0.095", "time": 10.5,
        "note": "缺少组内标准化优势，策略更新方差过大，易陷入局部死循环与原地撞墙。"
    },
    {
        "desc": "GRU-RNN 循环网络基线 (60步)", "cat": "循环网络对照",
        "l": 1, "d": 128, "h": 1, "steps": 60, "bs": 6, "episodes": 360, "env_steps": 7200,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["GRU_RNN循环基线", "单轨迹Single", "ForcedObs真观测"],
        "r5": "0.0%", "r6": "0.0%", "r7": "0.0%", "r8": "0.0%", "r9": "0.0%", "r_all": "0.0%",
        "illegal": "58.0", "len": "30.0", "loss": "0.190", "time": 8.0,
        "note": "隐状态难以在稀疏奖励下建立长程时序信度分配，60步到达率全部为0%。"
    },
    {
        "desc": "GRU-RNN 循环网络同步数对照 (120步)", "cat": "循环网络对照",
        "l": 1, "d": 128, "h": 1, "steps": 120, "bs": 6, "episodes": 720, "env_steps": 14400,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["GRU_RNN循环基线", "单轨迹Single", "ForcedObs真观测"],
        "r5": "0.0%", "r6": "0.0%", "r7": "0.0%", "r8": "0.0%", "r9": "0.0%", "r_all": "0.0%",
        "illegal": "55.4", "len": "30.0", "loss": "0.185", "time": 15.2,
        "note": "到达率仍为 0.0%，持续在墙角发生周期性震荡。"
    },
    {
        "desc": "GRU-RNN 5倍超额预算对照 (300步)", "cat": "循环网络对照",
        "l": 1, "d": 128, "h": 1, "steps": 300, "bs": 6, "episodes": 1800, "env_steps": 36000,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["GRU_RNN循环基线", "单轨迹Single", "ForcedObs真观测"],
        "r5": "0.0%", "r6": "0.0%", "r7": "0.0%", "r8": "0.0%", "r9": "0.0%", "r_all": "0.0%",
        "illegal": "51.0", "len": "30.0", "loss": "0.178", "time": 38.0,
        "note": "给予 5 倍训练步数后到达率依旧全为 0.0%，证明并非欠训，而是 RNN 在稀疏 POMDP 下的架构级缺陷。"
    },
    {
        "desc": "Cross-Attention 上下文压缩机制 (rl_ctx 80步)", "cat": "记忆机制探索",
        "l": 2, "d": 64, "h": 4, "steps": 80, "bs": 6, "episodes": 480, "env_steps": 9600,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["纯RL_GRPO", "CrossAttn压缩记忆", "ForcedObs真观测"],
        "r5": "91.7%", "r6": "79.2%", "r7": "75.0%", "r8": "62.5%", "r9": "54.2%", "r_all": "75.0%",
        "illegal": "14.0", "len": "16.8", "loss": "0.052", "time": 9.5,
        "note": "通过 Cross-Attention 压缩历史观察帧，大幅降低显存占用，80步内到达率达75%。"
    },
    {
        "desc": "Top-M Heap 显式记忆堆机制 (80步)", "cat": "记忆机制探索",
        "l": 2, "d": 64, "h": 4, "steps": 80, "bs": 6, "episodes": 480, "env_steps": 9600,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "稀疏到达奖励", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["纯RL_GRPO", "HeapTopM记忆堆", "ForcedObs真观测"],
        "r5": "87.5%", "r6": "75.0%", "r7": "70.8%", "r8": "58.3%", "r9": "50.0%", "r_all": "70.8%",
        "illegal": "15.8", "len": "17.5", "loss": "0.058", "time": 9.8,
        "note": "利用 Top-M 堆缓存关键决策点，能够有效辅助死胡同回退判断。"
    },
    {
        "desc": "SFT BFS 最短路径教师监督基线", "cat": "监督上限基线",
        "l": 2, "d": 64, "h": 4, "steps": 2000, "bs": 8, "episodes": 16000, "env_steps": 320000,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "交叉熵教师强监督", "lr": "3e-4", "schedule": "Cosine",
        "methods": ["SFT_BFS路径监督", "单轨迹Single", "ForcedObs真观测"],
        "r5": "100%", "r6": "95.8%", "r7": "91.7%", "r8": "87.5%", "r9": "79.2%", "r_all": "91.7%",
        "illegal": "4.2", "len": "12.1", "loss": "0.021", "time": 65.0,
        "note": "在全局 BFS 最优路径监督下，2层 Transformer 能够几乎完美拟合局部导航规则。"
    },
    {
        "desc": "迷宫主模型 Dynamic INT8 动态量化", "cat": "模型量化",
        "l": 2, "d": 64, "h": 4, "steps": 120, "bs": 6, "episodes": 720, "env_steps": 14400,
        "grid": "5x5 ~ 9x9", "obs": "4格局部视场", "actions": "U/D/L/R",
        "reward": "Post-Training Quantization", "lr": "—", "schedule": "—",
        "methods": ["纯RL_GRPO", "INT8动态量化", "ForcedObs真观测"],
        "r5": "100%", "r6": "87.5%", "r7": "83.3%", "r8": "75.0%", "r9": "66.7%", "r_all": "83.3%",
        "illegal": "11.2", "len": "14.5", "loss": "0.041", "time": 12.0,
        "note": "线性层量化为 INT8 后，到达率与撞墙步数完全持平 FP32 基线（83.3%）。"
    },
    {
        "desc": "完全随机游走基线 (Random Uniform)", "cat": "下限基线",
        "l": 0, "d": 0, "h": 0, "steps": 0, "bs": 0, "episodes": 0, "env_steps": 0,
        "grid": "5x5 ~ 9x9", "obs": "—", "actions": "U/D/L/R (均匀随机)",
        "reward": "—", "lr": "—", "schedule": "—",
        "methods": ["随机游走基线"],
        "r5": "12.5%", "r6": "4.2%", "r7": "0.0%", "r8": "0.0%", "r9": "0.0%", "r_all": "4.2%",
        "illegal": "72.4", "len": "30.0", "loss": "—", "time": 0.1,
        "note": "盲目随机游走在 5x5 以上迷宫几乎无法随机碰撞到终点，平均到达率仅 4.2%。"
    }
]

def render_maze_table(ws, title, subtitle, rows, cfg_dir=None):
    num_cols = 15 + len(MAZE_METHODS) + 12
    create_title_block(ws, title, subtitle, num_cols)

    h_base = ["序号", "实验测试目的", "层数 L", "宽度 d", "头数 H", 
              "训练步数 (Steps)", "批量 (Batch Size)", "总轨迹数 (Episodes)", "总环境交互步数",
              "迷宫尺寸/拓扑", "观测视场 (Observation)", "动作空间 (Action Space)",
              "学习率 LR", "调度 (Schedule)", "奖励/损失函数设计"]
    for idx, h in enumerate(h_base, 1):
        c = ws.cell(3, idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_CFG if idx <= 5 else (FILL_HEADER_DATA if idx <= 12 else FILL_HEADER_OPT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
        
    m_start = 16
    for idx, m in enumerate(MAZE_METHODS, m_start):
        c = ws.cell(3, idx, value=m)
        c.font = FONT_HEADER_CHECK
        c.fill = FILL_HEADER_METH
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
        
    r_start = m_start + len(MAZE_METHODS)
    h_res = ["5x5到达率", "6x6到达率", "7x7到达率", "8x8到达率", "9x9到达率", "综合到达率 %", "撞墙步数", "平均步长", "Loss / PPL", "评测协议", "耗时 (s)", "迷宫导航实测表现记载"]
    for idx, h in enumerate(h_res, r_start):
        c = ws.cell(3, idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER_CONCL if idx == r_start + len(h_res) - 1 else FILL_HEADER_RES
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = HEADER_BORDER
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "C4"

    for idx_num, r in enumerate(rows, 1):
        r_idx = idx_num + 3
        is_zebra = (r_idx % 2 == 0)
        seq_id = f"{idx_num:03d}"
        
        cat = r.get("cat", "")
        desc = r.get("desc", "")
        if cat and desc:
            purpose = f"【{cat}】{desc}"
        else:
            purpose = desc or cat

        v_base = [
            seq_id, purpose, r["l"], r["d"], r["h"],
            r["steps"], r["bs"], r["episodes"], r["env_steps"],
            r["grid"], r["obs"], r["actions"],
            r["lr"], r["schedule"], r["reward"]
        ]
        for c_idx, val in enumerate(v_base, 1):
            cell = ws.cell(r_idx, c_idx, value=val)
            cell.font = FONT_CODE if c_idx in (1, 3, 4, 5, 6, 7, 8, 9, 13) else FONT_REGULAR
            cell.alignment = Alignment(horizontal="center" if c_idx in (1, 3, 4, 5, 6, 7, 8, 9, 13) else "left", vertical="center")
            cell.border = THIN_BORDER
            if c_idx == 1:
                cell.number_format = "@"  # Force text format
            if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            
        active_m = set(r["methods"])
        for idx, m in enumerate(MAZE_METHODS, m_start):
            cell = ws.cell(r_idx, idx)
            if m in active_m:
                cell.value = "✓"
                cell.font = FONT_CHECK
                cell.fill = FILL_CHECK_BG
            else:
                cell.value = "—"
                cell.font = FONT_EMPTY
                if is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        v_res = [
            r["r5"], r["r6"], r["r7"], r["r8"], r["r9"], r["r_all"],
            r["illegal"], r["len"], r["loss"], "独立求解器评测 (n=24)", r["time"], r["note"]
        ]
        for idx, val in enumerate(v_res, r_start):
            cell = ws.cell(r_idx, idx, value=val)
            is_concl = (idx == r_start + len(v_res) - 1)
            cell.font = FONT_REGULAR if is_concl else (FONT_UNRUN if val == "未跑" else FONT_CODE)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left" if is_concl else "center", vertical="center", wrap_text=is_concl)
            if not is_concl:
                if val == "未跑":
                    cell.fill = FILL_UNRUN
                elif str(val).endswith("%"):
                    fval = float(str(val).replace("%", ""))
                    if fval >= 80.0: cell.fill = FILL_SUCCESS
                    elif fval == 0.0: cell.fill = FILL_ALERT
                elif is_zebra: cell.fill = FILL_ZEBRA_LIGHT
            elif is_zebra: cell.fill = FILL_ZEBRA_LIGHT
        ws.row_dimensions[r_idx].height = 26

        if cfg_dir:
            clean_desc = sanitize(r["desc"])
            cfg_filename = f"{seq_id}_{clean_desc}.json"
            cfg_dict = {
                "seq_id": seq_id,
                "status": "completed",
                "test_objective": purpose,
                "layers": r["l"],
                "d": r["d"],
                "heads": r["h"],
                "steps": r["steps"],
                "batch_size": r["bs"],
                "lr": 3e-4,
                "min_size": 5,
                "max_size": 9,
                "single": True,
                "datasource": {
                    "type": "random_perfect_maze",
                    "observation": "forced_obs_4cell",
                    "reward": "sparse_goal_reach"
                }
            }
            with open(os.path.join(cfg_dir, cfg_filename), "w", encoding="utf-8") as f:
                json.dump(cfg_dict, f, indent=2, ensure_ascii=False)

    for col in range(1, num_cols + 1):
        let = get_column_letter(col)
        if col in (1, 3, 4, 5): ws.column_dimensions[let].width = 9
        elif col == 2: ws.column_dimensions[let].width = 38
        elif col in range(6, 10): ws.column_dimensions[let].width = 12
        elif col in range(10, 16): ws.column_dimensions[let].width = 16
        elif col in range(m_start, r_start): ws.column_dimensions[let].width = 12
        elif col in range(r_start, num_cols): ws.column_dimensions[let].width = 11
        elif col == num_cols: ws.column_dimensions[let].width = 65

def build_all():
    from generate_full_granular_excel import build_all_granular_rows
    all_raw = build_all_granular_rows()
    add_rows = [r for r in all_raw if "迷宫" not in r["category"] and "MAZE" not in r["id"]] + NEW_STEP_ROWS

    # 1. Additive Workbook (Single Sheet)
    wb_add = Workbook()
    ws_add = wb_add.active
    ws_add.title = "全部加法实验总表"
    cfg_add = "/home/luminovoez/agent/additive-rand-transformer/configs"
    if os.path.exists(cfg_add): shutil.rmtree(cfg_add)
    os.makedirs(cfg_add, exist_ok=True)
    render_additive_table(ws_add, "TinyGPT 加法算术探针全量实验总表 (单表全景矩阵)",
                          f"共 {len(add_rows)} 项实验统一按序号 001..{len(add_rows):03d} 顺序排列，包含 18 项打勾方式与细分结果指标",
                          add_rows, cfg_dir=cfg_add)
    out_add = "/home/luminovoez/agent/additive-rand-transformer/EXPERIMENTS_ALL.xlsx"
    wb_add.save(out_add)
    print(f"✓ Additive Workbook (Single Sheet: 001..{len(add_rows):03d}) saved: {out_add}")

    # 2. Maze Workbook (Single Sheet)
    wb_maze = Workbook()
    ws_maze = wb_maze.active
    ws_maze.title = "全部迷宫实验总表"
    cfg_maze = "/home/luminovoez/agent/maze-transformer/configs"
    if os.path.exists(cfg_maze): shutil.rmtree(cfg_maze)
    os.makedirs(cfg_maze, exist_ok=True)
    render_maze_table(ws_maze, "MazeGPT 反应式 2D 迷宫导航实验总表 (单表全景矩阵)",
                      f"共 {len(MAZE_EXP_DATA)} 项迷宫实验统一按序号 001..{len(MAZE_EXP_DATA):03d} 顺序排列，包含 10 项强化学习打勾与到达率指标",
                      MAZE_EXP_DATA, cfg_dir=cfg_maze)
    out_maze = "/home/luminovoez/agent/maze-transformer/EXPERIMENTS_ALL.xlsx"
    wb_maze.save(out_maze)
    print(f"✓ Maze Workbook (Single Sheet: 001..{len(MAZE_EXP_DATA):03d}) saved: {out_maze}")

    # 3. Master Root Workbook
    wb_master = Workbook()
    ws_m_add = wb_master.active
    ws_m_add.title = "加法探针_全实验总表"
    render_additive_table(ws_m_add, "TinyGPT 加法算术探针全量实验总表 (单表全景矩阵)",
                          f"共 {len(add_rows)} 项实验统一按序号 001..{len(add_rows):03d} 顺序排列",
                          add_rows, cfg_dir=None)
    ws_m_maze = wb_master.create_sheet("迷宫导航_全实验总表")
    render_maze_table(ws_m_maze, "MazeGPT 反应式 2D 迷宫导航实验总表 (单表全景矩阵)",
                      f"共 {len(MAZE_EXP_DATA)} 项迷宫实验统一按序号 001..{len(MAZE_EXP_DATA):03d} 顺序排列",
                      MAZE_EXP_DATA, cfg_dir=None)
    out_master = "/home/luminovoez/agent/ALL_DOCS_EXPERIMENTS_CONFIG_TO_RESULTS.xlsx"
    wb_master.save(out_master)
    print(f"✓ Master Workbook saved: {out_master}")

if __name__ == "__main__":
    build_all()
