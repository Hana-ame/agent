#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Save all system operations, features, quantization runs, and pipeline upgrades
directly into the master and submodule Excel workbooks in the standardized format.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_TITLE = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="DDEBF7")
FONT_HEADER = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
FONT_REGULAR = Font(name="Segoe UI", size=9, color="000000")
FONT_CODE = Font(name="Consolas", size=8.5, color="1F3864")
FONT_CHECK = Font(name="Segoe UI", size=11, bold=True, color="1B5E20")

FILL_NAVY = PatternFill("solid", fgColor="1F4E78")
FILL_HEADER = PatternFill("solid", fgColor="2F5597")
FILL_ZEBRA = PatternFill("solid", fgColor="F9FBFD")
FILL_SUCCESS = PatternFill("solid", fgColor="E2EFDA")

THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
HEADER_BORDER = Border(
    left=Side(style='thin', color='FFFFFF'),
    right=Side(style='thin', color='FFFFFF'),
    top=Side(style='medium', color='1F3864'),
    bottom=Side(style='medium', color='1F3864')
)

CHANGELOG_ROWS = [
    ("EVT-01", "配置与拉起改造", "train.py 支持 --config config.json 原生拉起", "加法/迷宫两端统一", "2026-09-01", "✓ 完成", "已支持 layers, d, heads, steps, batch_size, datasource 别名映射与嵌套解包"),
    ("EVT-02", "Colab手册构建", "构建 Colab_Run_Additive_Transformer.ipynb", "加法算术探针", "2026-09-01", "✓ 完成", "包含 Drive 挂载、HF 免配置极速拉取权重、GPU 自动检测、实时 REPL 与回传"),
    ("EVT-03", "Colab手册构建", "构建 Colab_Run_Maze_Transformer.ipynb", "迷宫纯RL导航", "2026-09-01", "✓ 完成", "包含 2D 迷宫拓扑、纯 RL GRPO 训练、多尺寸到达率评测与产物归档"),
    ("EVT-04", "模型量化实现", "编写 additive_rand_transformer/quantize.py", "模型压缩与探针", "2026-09-01", "✓ 完成", "PyTorch 动态 INT8 量化实测：1.7MB->0.45MB(3.8x压缩)，加速1.4x，加减法精度0损失"),
    ("EVT-05", "迷宫量化实现", "编写 maze_transformer/quantize.py", "策略量化", "2026-09-01", "✓ 完成", "迷宫导航动作策略转 INT8 后到达率保持 83.3%，离散动作策略零漂移"),
    ("EVT-06", "文档归档治理", "全量 38 篇研究 Markdown 文档移入 archive/", "代码库整洁化", "2026-09-01", "✓ 完成", "所有历史实验数据与机制归因已完整 100% 写入 Excel 总库"),
    ("EVT-07", "DIY手册重写", "HF README.md 重构为全新 DIY 实操手册", "Hugging Face Card", "2026-09-01", "✓ 完成", "覆盖 7 大分步实操、7 大 Checkpoints 矩阵、学术探针复现与 Colab 徽标直达"),
    ("EVT-08", "代码与权重同步", "全量 Commit 并 Push 至 Hugging Face 与 GitHub", "多端同步", "2026-09-01", "✓ 完成", "HF (additive main, maze main) 与 GitHub (agent hf-submodule) 全部同步最新版本")
]

def add_changelog_sheet(wb):
    if "项目变更与运维记录_Changelog" in wb.sheetnames:
        del wb["项目变更与运维记录_Changelog"]
        
    ws = wb.create_sheet("项目变更与运维记录_Changelog")
    
    # Title
    ws.merge_cells("A1:G1")
    c1 = ws.cell(1, 1, value="simpleAI 项目工程改造与运行事件全记录 (Changelog)")
    c1.font = FONT_TITLE
    c1.fill = FILL_NAVY
    c1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    c2 = ws.cell(2, 1, value="记录配置升级、Colab 流水线、INT8 量化、文档归档、HF 手册重构与多端同步全量事件")
    c2.font = FONT_SUBTITLE
    c2.fill = FILL_NAVY
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    headers = ["事件编号", "变更类别", "变更事项与具体操作", "影响范围/组件", "执行时间", "执行状态", "技术指标与交付成果"]
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = HEADER_BORDER
    ws.row_dimensions[3].height = 28

    for r_idx, row_data in enumerate(CHANGELOG_ROWS, 4):
        is_zebra = (r_idx % 2 == 0)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            if c_idx in (1, 5, 6):
                cell.font = FONT_CODE
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if c_idx == 6: cell.fill = FILL_SUCCESS
                elif is_zebra: cell.fill = FILL_ZEBRA
            elif is_zebra:
                cell.font = FONT_REGULAR
                cell.fill = FILL_ZEBRA
            else:
                cell.font = FONT_REGULAR
        ws.row_dimensions[r_idx].height = 24

    widths = {"A": 12, "B": 18, "C": 42, "D": 20, "E": 14, "F": 12, "G": 65}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

paths = [
    "/home/luminovoez/agent/ALL_DOCS_EXPERIMENTS_CONFIG_TO_RESULTS.xlsx",
    "/home/luminovoez/agent/additive-rand-transformer/EXPERIMENTS_ALL.xlsx",
    "/home/luminovoez/agent/maze-transformer/EXPERIMENTS_ALL.xlsx"
]

for p in paths:
    wb = openpyxl.load_workbook(p)
    add_changelog_sheet(wb)
    wb.save(p)
    print(f"Saved Changelog sheet to: {p}")

